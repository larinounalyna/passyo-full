"""
BIM File Parser - Extract components from various formats
Supports: Excel, CSV, IFC
"""

import csv
from decimal import Decimal
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)

# ============================================================================
# COMPONENT MAPPING - Estimate labor hours based on component type
# ============================================================================

LABOR_HOUR_ESTIMATES = {
    'IfcWall': Decimal('2.0'),
    'IfcWallStandardCase': Decimal('2.0'),
    'IfcCurtainWall': Decimal('3.0'),
    'IfcSlab': Decimal('1.5'),
    'IfcSlabStandardCase': Decimal('1.5'),
    'IfcColumn': Decimal('3.0'),
    'IfcBeam': Decimal('2.5'),
    'IfcFooting': Decimal('2.5'),
    'IfcPile': Decimal('4.0'),
    'IfcRoof': Decimal('4.0'),
    'IfcWindow': Decimal('1.5'),
    'IfcDoor': Decimal('1.0'),
    'IfcStair': Decimal('3.5'),
    'IfcStaircase': Decimal('3.5'),
    'IfcStairFlight': Decimal('3.0'),
    'IfcRamp': Decimal('2.0'),
    'IfcRailing': Decimal('1.5'),
    'IfcCovering': Decimal('1.0'),
    'IfcPlate': Decimal('1.5'),
    'IfcMember': Decimal('2.0'),
    'IfcFurnishingElement': Decimal('0.5'),
}

MATERIAL_CODE_MAPPING = {
    'concrete': 'CONCRETE_MIX',
    'steel': 'REBAR',
    'brick': 'MASONRY',
    'wood': 'TIMBER',
    'aluminum': 'ALUMINUM',
    'glass': 'GLASS',
    'excavation': 'SOIL_EXCAVATION',
    'foundation': 'FOUNDATION',
}

# ============================================================================
# SMART COLUMN MAPPER - Recognizes column names regardless of format
# ============================================================================

import re

# Each field maps to keywords (scored by priority) that identify it.
# A column header is normalized (lowercased, punctuation removed) then
# matched against these keywords. The field with the best score wins.
FIELD_KEYWORDS = {
    'component': {
        'must_have': ['component', 'name', 'item', 'element', 'designation',
                      'titre', 'composant', 'bezeichnung', 'artikel'],
        'boost': ['desc', 'label', 'title'],
        'exclude': ['type', 'code', 'price', 'cost', 'rate', 'hour', 'qty',
                     'quantity', 'unit', 'material'],
    },
    'description': {
        'must_have': ['description', 'desc', 'detail', 'notes', 'remark',
                      'specification', 'spec'],
        'boost': ['info', 'comment', 'text'],
        'exclude': ['type', 'code', 'price', 'cost', 'rate', 'hour'],
    },
    'ifc_type': {
        'must_have': ['type', 'ifc', 'category', 'class', 'classification',
                      'element type', 'categorie', 'klasse'],
        'boost': ['kind', 'group', 'family'],
        'exclude': ['price', 'cost', 'rate', 'hour', 'material', 'soil',
                     'concrete', 'equipment'],
    },
    'quantity': {
        'must_have': ['quantity', 'qty', 'amount', 'count', 'number', 'vol',
                      'volume', 'area', 'length', 'quantite', 'menge', 'anzahl'],
        'boost': ['total', 'net', 'gross'],
        'exclude': ['price', 'cost', 'rate', 'unit', 'labor', 'hour'],
    },
    'unit': {
        'must_have': ['unit', 'uom', 'measure', 'unite', 'einheit'],
        'boost': ['measurement'],
        'exclude': ['price', 'cost', 'rate', 'quantity', 'qty'],
    },
    'price': {
        'must_have': ['price', 'cost', 'rate', 'amount', 'value', 'prix',
                      'cout', 'preis', 'kosten', 'tarif'],
        'boost': ['unit', 'material', 'base', 'per', 'total'],
        'exclude': ['labor', 'labour', 'work', 'hour', 'wage', 'salary',
                     'equipment', 'rate per hour'],
    },
    'labor_hours': {
        'must_have': ['hour', 'hrs', 'time', 'duration', 'heure', 'stunden',
                      'heures'],
        'boost': ['labor', 'labour', 'work', 'man', 'effort'],
        'exclude': ['rate', 'price', 'cost', 'wage', 'salary', 'per hour'],
    },
    'labor_rate': {
        'must_have': ['wage', 'salary', 'hourly'],
        'boost': ['labor rate', 'labour rate', 'rate per hour', 'pay',
                  'cost per hour'],
        'exclude': [],
    },
    'material_code': {
        'must_have': ['material', 'code', 'mat', 'product', 'materiau',
                      'werkstoff'],
        'boost': ['id', 'ref', 'reference', 'sku', 'catalog'],
        'exclude': ['price', 'cost', 'rate', 'hour', 'labor', 'labour',
                     'description', 'name'],
    },
}

# Special compound detection: if a header contains BOTH a boost word and a
# must_have word for the WRONG field, we can resolve ambiguity. E.g.
# "labor rate" has "rate" (price must_have) but "labor" (price exclude).


def _normalize(text: str) -> str:
    """Normalize a header: lowercase, remove punctuation, collapse whitespace"""
    text = str(text).lower().strip()
    text = re.sub(r'[_\-/\\.,;:()#\[\]{}]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _score_field(normalized_header: str, field: str) -> float:
    """Score how well a normalized header matches a field definition"""
    rules = FIELD_KEYWORDS.get(field, {})
    score = 0.0

    # Check exclusions first - strong negative signal
    for exc in rules.get('exclude', []):
        if exc in normalized_header:
            score -= 5.0

    # Check must_have keywords
    for kw in rules.get('must_have', []):
        if kw in normalized_header:
            score += 10.0

    # Check boost keywords
    for kw in rules.get('boost', []):
        if kw in normalized_header:
            score += 3.0

    return score


def _resolve_labor_ambiguity(normalized_header: str, current_field: str) -> str:
    """Resolve ambiguity between labor_hours, labor_rate, and price fields"""
    has_labor = any(w in normalized_header for w in ['labor', 'labour', 'work', 'man'])
    has_rate = any(w in normalized_header for w in ['rate', 'wage', 'salary', 'pay', 'hourly', 'per hour'])
    has_hour = any(w in normalized_header for w in ['hour', 'hrs', 'time', 'duration'])
    has_price = any(w in normalized_header for w in ['price', 'cost', 'prix'])
    has_material = any(w in normalized_header for w in ['material', 'mat'])

    # "labor rate" or "wage" -> labor_rate
    if has_labor and has_rate:
        return 'labor_rate'
    # "labor hours" or "man hours" -> labor_hours
    if has_labor and has_hour:
        return 'labor_hours'
    # "unit price" or "material cost" -> price
    if has_price or (has_material and (has_rate or 'cost' in normalized_header)):
        return 'price'
    # "rate" alone with no other context -> could be labor_rate
    if has_rate and not has_price and not has_material:
        return 'labor_rate'

    return current_field


def map_columns(headers: List[str]) -> Dict[str, str]:
    """
    Smart column mapper: takes a list of raw column headers and returns
    a mapping of {field_name: original_header_name}.

    Uses keyword scoring with ambiguity resolution. Works with any language
    or naming convention (snake_case, camelCase, spaces, French, German, etc.)
    """
    normalized = {h: _normalize(h) for h in headers if h}

    # Score every header against every field
    scores = {}  # {original_header: {field: score}}
    for orig, norm in normalized.items():
        scores[orig] = {}
        for field in FIELD_KEYWORDS:
            scores[orig][field] = _score_field(norm, field)

    # Assign fields greedily: highest score first, no field assigned twice
    mapping = {}
    assigned_headers = set()
    assigned_fields = set()

    # Build list of (score, header, field) and sort descending
    candidates = []
    for orig in scores:
        for field, score in scores[orig].items():
            if score > 0:
                candidates.append((score, orig, field))
    candidates.sort(key=lambda x: -x[0])

    for score, header, field in candidates:
        if header in assigned_headers or field in assigned_fields:
            continue

        # Apply ambiguity resolution
        norm = normalized[header]
        resolved_field = _resolve_labor_ambiguity(norm, field)
        if resolved_field != field:
            if resolved_field not in assigned_fields:
                field = resolved_field
            else:
                continue

        mapping[field] = header
        assigned_headers.add(header)
        assigned_fields.add(field)

    # If 'component' not found but 'description' is, use description as component name
    if 'component' not in mapping and 'description' in mapping:
        mapping['component'] = mapping['description']

    logger.info(f"Smart column mapping: {mapping}")
    return mapping


def _get_mapped_value(row_dict: Dict, mapping: Dict[str, str], field: str, default=None):
    """Get a value from row_dict using the smart column mapping"""
    header = mapping.get(field)
    if header and header in row_dict:
        val = row_dict[header]
        if val is not None and str(val).strip() != '':
            return val
    return default


# ============================================================================
# EXCEL PARSER
# ============================================================================

def _build_component(idx: int, row_dict: Dict, mapping: Dict[str, str], source: str) -> Dict[str, Any]:
    """Build a component dict from a row using the smart column mapping"""
    component_name = _get_mapped_value(row_dict, mapping, 'component',
                     default=_get_mapped_value(row_dict, mapping, 'description',
                     default=f"Component_{idx+1}"))

    ifc_type = _get_mapped_value(row_dict, mapping, 'ifc_type', default='Unknown')

    try:
        quantity = float(_get_mapped_value(row_dict, mapping, 'quantity', default=1.0))
    except (ValueError, TypeError):
        quantity = 1.0

    unit = _get_mapped_value(row_dict, mapping, 'unit', default='m3')

    material_code = _get_mapped_value(row_dict, mapping, 'material_code',
                                      default=f"MAT_{idx+1}")

    try:
        price = float(_get_mapped_value(row_dict, mapping, 'price', default=0))
    except (ValueError, TypeError):
        price = 0.0

    try:
        labor_hours = float(_get_mapped_value(row_dict, mapping, 'labor_hours', default=0))
    except (ValueError, TypeError):
        labor_hours = 0
    if labor_hours == 0:
        labor_hours = float(LABOR_HOUR_ESTIMATES.get(str(ifc_type), Decimal('1.0')))

    try:
        labor_rate = float(_get_mapped_value(row_dict, mapping, 'labor_rate', default=30.0))
    except (ValueError, TypeError):
        labor_rate = 30.0

    return {
        "id": f"{source}_comp_{idx+1:04d}",
        "ifc_type": str(ifc_type),
        "description": str(component_name),
        "quantity": float(quantity),
        "unit": str(unit),
        "material_code": str(material_code),
        "base_material_price": price,
        "labor_hours_base": labor_hours,
        "labor_rate_per_hour": labor_rate,
        "source": source,
    }


def parse_excel_bom(filepath: str) -> List[Dict[str, Any]]:
    """Parse Bill of Materials from Excel file using openpyxl directly"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        components = []

        rows = ws.iter_rows()
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError("Excel file is empty")

        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
        mapping = map_columns(headers)

        for idx, row in enumerate(rows):
            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_dict[headers[col_idx]] = cell.value

            if all(v is None for v in row_dict.values()):
                continue

            components.append(_build_component(idx, row_dict, mapping, "excel"))

        wb.close()
        logger.info(f"Parsed {len(components)} components from Excel")
        return components

    except ImportError:
        raise
    except Exception as e:
        logger.error(f"Error parsing Excel: {str(e)}")
        raise ValueError(f"Failed to parse Excel file: {str(e)}")


# ============================================================================
# CSV PARSER
# ============================================================================

def parse_csv_bom(filepath: str) -> List[Dict[str, Any]]:
    """Parse Bill of Materials from CSV file"""
    try:
        components = []

        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            mapping = map_columns(headers)

            for idx, row in enumerate(reader):
                components.append(_build_component(idx, row, mapping, "csv"))

        logger.info(f"Parsed {len(components)} components from CSV")
        return components

    except Exception as e:
        logger.error(f"Error parsing CSV: {str(e)}")
        raise ValueError(f"Failed to parse CSV file: {str(e)}")


# ============================================================================
# IFC PARSER (requires ifcopenshell)
# ============================================================================

# All supported IFC element types
IFC_ELEMENT_TYPES = [
    'IfcWall', 'IfcWallStandardCase',
    'IfcColumn', 'IfcBeam', 'IfcMember',
    'IfcSlab', 'IfcSlabStandardCase',
    'IfcFooting', 'IfcPile',
    'IfcRoof',
    'IfcDoor', 'IfcWindow',
    'IfcStair', 'IfcStairFlight', 'IfcStaircase',
    'IfcRamp', 'IfcRailing',
    'IfcPlate', 'IfcCovering',
    'IfcCurtainWall',
    'IfcFurnishingElement', 'IfcEquipmentElement',
]


def _try_extract_quantity_from_qto(entity) -> tuple:
    """Try to extract quantity from Qto_ property sets (IfcElementQuantity)"""
    quantity = None
    unit = None

    try:
        if not hasattr(entity, 'IsDefinedBy'):
            return None, None

        for rel in entity.IsDefinedBy:
            if not hasattr(rel, 'RelatingPropertyDefinition'):
                continue

            prop_def = rel.RelatingPropertyDefinition

            # Check for IfcElementQuantity (Qto_* property sets)
            if prop_def.is_a('IfcElementQuantity'):
                for q in prop_def.Quantities:
                    q_type = q.is_a()
                    if q_type == 'IfcQuantityVolume':
                        quantity = float(q.VolumeValue)
                        unit = 'm3'
                        return quantity, unit
                    elif q_type == 'IfcQuantityArea':
                        if quantity is None:  # Volume takes priority
                            quantity = float(q.AreaValue)
                            unit = 'm2'
                    elif q_type == 'IfcQuantityLength':
                        if quantity is None and unit is None:
                            quantity = float(q.LengthValue)
                            unit = 'm'
                    elif q_type == 'IfcQuantityWeight':
                        if quantity is None and unit is None:
                            quantity = float(q.WeightValue)
                            unit = 'kg'
                    elif q_type == 'IfcQuantityCount':
                        if quantity is None and unit is None:
                            quantity = float(q.CountValue)
                            unit = 'each'
    except Exception as ex:
        logger.debug(f"Qto extraction failed: {ex}")

    return quantity, unit


def _try_extract_quantity_from_properties(entity) -> tuple:
    """Fallback: extract quantity from regular property sets"""
    quantity = None
    unit = None

    try:
        if not hasattr(entity, 'IsDefinedBy'):
            return None, None

        for definition in entity.IsDefinedBy:
            if not hasattr(definition, 'RelatingPropertyDefinition'):
                continue
            prop_def = definition.RelatingPropertyDefinition
            if not hasattr(prop_def, 'HasProperties'):
                continue
            for prop in prop_def.HasProperties:
                if not (hasattr(prop, 'Name') and hasattr(prop, 'NominalValue')):
                    continue
                try:
                    nom_val = prop.NominalValue
                    if not hasattr(nom_val, 'wrappedValue'):
                        continue
                    val = float(nom_val.wrappedValue)
                    prop_name = str(prop.Name)
                    if 'Volume' in prop_name:
                        quantity = val
                        unit = 'm3'
                    elif 'Area' in prop_name and quantity is None:
                        quantity = val
                        unit = 'm2'
                    elif 'Length' in prop_name and quantity is None:
                        quantity = val
                        unit = 'm'
                except (ValueError, TypeError):
                    pass
    except Exception as ex:
        logger.debug(f"Property extraction failed: {ex}")

    return quantity, unit


def _extract_material(entity) -> str:
    """Extract material information from IFC entity"""
    try:
        if not hasattr(entity, 'HasAssociations'):
            return "UNKNOWN"

        for assoc in entity.HasAssociations:
            if not hasattr(assoc, 'RelatingMaterial'):
                continue
            material = assoc.RelatingMaterial

            # Direct material
            if hasattr(material, 'Name') and material.Name:
                return str(material.Name)

            # IfcMaterialLayerSetUsage -> IfcMaterialLayerSet -> layers
            if material.is_a('IfcMaterialLayerSetUsage'):
                layer_set = material.ForLayerSet
                if layer_set and layer_set.MaterialLayers:
                    names = []
                    for layer in layer_set.MaterialLayers:
                        if layer.Material and layer.Material.Name:
                            names.append(str(layer.Material.Name))
                    if names:
                        return " / ".join(names)

            # IfcMaterialLayerSet directly
            if material.is_a('IfcMaterialLayerSet'):
                if material.MaterialLayers:
                    names = []
                    for layer in material.MaterialLayers:
                        if layer.Material and layer.Material.Name:
                            names.append(str(layer.Material.Name))
                    if names:
                        return " / ".join(names)

            # IfcMaterialList
            if material.is_a('IfcMaterialList'):
                if material.Materials:
                    names = [str(m.Name) for m in material.Materials if m.Name]
                    if names:
                        return " / ".join(names)

    except Exception as ex:
        logger.debug(f"Material extraction failed: {ex}")

    return "UNKNOWN"


def parse_ifc_bom(filepath: str) -> List[Dict[str, Any]]:
    """Parse Bill of Materials from IFC BIM file"""
    try:
        import ifcopenshell
    except ImportError:
        raise ImportError("ifcopenshell not installed. Run: pip install ifcopenshell")

    try:
        ifc_file = ifcopenshell.open(filepath)
        components = []
        parsed_entities = set()

        for elem_type in IFC_ELEMENT_TYPES:
            try:
                for entity in ifc_file.by_type(elem_type):
                    entity_id = getattr(entity, 'GlobalId', str(entity))
                    if entity_id in parsed_entities:
                        continue
                    parsed_entities.add(entity_id)

                    ifc_type = entity.is_a()
                    component_id = entity.GlobalId if hasattr(entity, 'GlobalId') else f"entity_{len(components)}"
                    component_name = entity.Name if hasattr(entity, 'Name') and entity.Name else ifc_type

                    # Try Qto first, then fallback to properties
                    quantity, unit = _try_extract_quantity_from_qto(entity)
                    if quantity is None:
                        quantity, unit = _try_extract_quantity_from_properties(entity)
                    if quantity is None:
                        quantity = 1.0
                        unit = "each"

                    # Get material
                    material_code = _extract_material(entity)

                    # Estimate labor hours
                    labor_hours = float(LABOR_HOUR_ESTIMATES.get(ifc_type, Decimal('1.0')))

                    component = {
                        "id": str(component_id),
                        "ifc_type": ifc_type,
                        "description": str(component_name),
                        "quantity": quantity,
                        "unit": unit,
                        "material_code": material_code,
                        "base_material_price": 0,
                        "labor_hours_base": labor_hours,
                        "labor_rate_per_hour": 30.0,
                        "source": "ifc",
                        "needs_pricing": True,
                    }

                    components.append(component)

            except Exception as ex:
                logger.debug(f"Could not extract {elem_type}: {ex}")
                continue

        if not components:
            raise ValueError("IFC file loaded but no building elements could be extracted.")

        logger.info(f"Parsed {len(components)} components from IFC file")
        return components

    except ImportError:
        raise
    except Exception as e:
        logger.error(f"Error parsing IFC: {str(e)}")
        raise ValueError(f"Failed to parse IFC file: {str(e)}")


# ============================================================================
# AUTO-DETECT FILE FORMAT AND PARSE
# ============================================================================

def parse_bim_file(filepath: str) -> List[Dict[str, Any]]:
    """Auto-detect BIM file format and parse accordingly"""
    if filepath.endswith(('.xlsx', '.xls')):
        return parse_excel_bom(filepath)
    elif filepath.endswith('.csv'):
        return parse_csv_bom(filepath)
    elif filepath.endswith('.ifc'):
        return parse_ifc_bom(filepath)
    else:
        raise ValueError(
            f"Unsupported file format: {filepath}. "
            "Supported: .xlsx, .xls, .csv, .ifc"
        )


# ============================================================================
# VALIDATION
# ============================================================================

def validate_components(components: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Validate parsed components and identify missing data"""
    issues = {
        "missing_prices": [],
        "missing_labor_hours": [],
        "invalid_quantities": [],
        "total_components": len(components),
    }

    for comp in components:
        if comp.get('base_material_price', 0) == 0:
            issues['missing_prices'].append(comp['id'])

        if comp.get('labor_hours_base', 0) == 0:
            issues['missing_labor_hours'].append(comp['id'])

        try:
            if float(comp.get('quantity', 0)) <= 0:
                issues['invalid_quantities'].append(comp['id'])
        except (ValueError, TypeError):
            issues['invalid_quantities'].append(comp['id'])

    issues['status'] = 'valid' if not any([
        issues['missing_prices'],
        issues['missing_labor_hours'],
        issues['invalid_quantities'],
    ]) else 'has_issues'

    return issues


if __name__ == "__main__":
    print("BIM Parser module loaded successfully")
    print("\nSupported formats: Excel (.xlsx, .xls), CSV (.csv), IFC (.ifc)")
